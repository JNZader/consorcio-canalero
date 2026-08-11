"""Integration coverage for slice 3b's xlsx export (real PG).

design.md D7: ``GET /rainfall/analyses/{revision}.xlsx`` beside the CSV route,
under the same router-level ``require_admin_or_operator``. Two sheets --
**Resumen** (the same ``metric_rows(normalize_snapshot(...))`` projection the
audit CSV serializes, plus the disclosures a downloaded file needs because it
outlives the screen) and **Serie diaria** (the slice-3a series module).

The load-bearing property, and the reason both sheets are built from the SAME
two calls the JSON routes make rather than from re-derived numbers: a workbook
that quietly disagrees with the analysis it claims to export is worse than no
workbook at all, because it is the artifact that gets forwarded, printed and
argued over months later. So the parity test below compares two SERVED
artifacts (the xlsx against the audit CSV and against the ``/series`` JSON),
not the xlsx against a hand-built expectation.

Lluvia insights slice 3b: xlsx export + TS series/xlsx contract (D7).
"""

import csv
from datetime import UTC, date, datetime, timedelta
from io import BytesIO, StringIO

import pytest
from openpyxl import load_workbook

from app.domains.geo.rainfall.adapters.gee_client import BASELINE_ASSET_VERSION, asset_name_for

# The 3a fixture builders, imported rather than re-copied ON PURPOSE (repo
# precedent: `test_ficha_canal_cuenca.py` imports `test_generate_canal_catchments`).
# This file's central claim is that the xlsx reads the same envelope and the
# same series the JSON does; a second, hand-copied fixture would let the two
# drift apart and the parity assertions would then be measuring the fixtures.
from tests.new.geo.rainfall.test_rainfall_series_consistency import (
    _CURVE_CUTOFF,
    _CURVE_YEAR,
    _ZONE_FAMILY,
    _build_revision,
    _daily_rows,
    _days_through,
    _fake_revision,
    _persist_zone_rows,
    _seed_full_baseline,
)


def _client(db):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.auth import require_admin_or_operator
    from app.db.session import get_db
    from app.domains.geo.rainfall.router import router

    app = FastAPI()
    app.dependency_overrides[require_admin_or_operator] = lambda: None
    app.dependency_overrides[get_db] = lambda: db
    app.include_router(router)
    return TestClient(app)


def _seeded_revision(db, *, scope_id: str, baseline: bool = True):
    """One revision built through the real production path, with a full
    1991-2020 baseline so `annual.normal`/`annual.percentile` are served and
    the normal curve exists."""
    asset = asset_name_for("zone", scope_id)
    now = datetime(_CURVE_YEAR, 3, 2, 12, 0, tzinfo=UTC)
    if baseline:
        _seed_full_baseline(db, asset=asset, cutoff=_CURVE_CUTOFF)
    _persist_zone_rows(
        db,
        scope_id=scope_id,
        rows=_daily_rows(
            date(_CURVE_YEAR, 1, 1),
            _days_through(_CURVE_CUTOFF, _CURVE_YEAR),
            3.0,
            provider_revision=_ZONE_FAMILY,
        ),
    )
    return _build_revision(db, scope_id=scope_id, year=_CURVE_YEAR, now=now)


def _sheets(content: bytes):
    return load_workbook(BytesIO(content), read_only=True, data_only=True)


def _rows(sheet) -> list[list]:
    """Every row padded to the widest one.

    A read-only worksheet yields each row at its own width -- a label/value
    row is 2 cells, the spacer row is 0 -- so without the pad every positional
    read here would be an `IndexError` on exactly the rows this file cares
    about (the disclosure stamps and the empty "Normal acumulada" column).
    """
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    width = max((len(row) for row in rows), default=0)
    return [row + [None] * (width - len(row)) for row in rows]


def _labelled(rows: list[list]) -> dict[str, object]:
    """The Resumen header block, as {label: value} — every row before the
    metric table whose first cell is a non-empty string."""
    from app.domains.geo.rainfall.export import METRIC_TABLE_HEADER

    block: dict[str, object] = {}
    for row in rows:
        if list(row[: len(METRIC_TABLE_HEADER)]) == list(METRIC_TABLE_HEADER):
            break
        if row and isinstance(row[0], str) and row[0]:
            block[row[0]] = row[1] if len(row) > 1 else None
    return block


def _metric_table(rows: list[list]) -> list[dict]:
    """The Resumen metric rows, keyed by the header labels."""
    from app.domains.geo.rainfall.export import METRIC_TABLE_HEADER

    header = list(METRIC_TABLE_HEADER)
    start = next(index for index, row in enumerate(rows) if list(row[: len(header)]) == header)
    return [
        dict(zip(header, row, strict=False))
        for row in rows[start + 1 :]
        if any(cell is not None for cell in row)
    ]


# ---------------------------------------------------------------------------
# 3b.1 / 3b.2: the two sheets, and the authorization boundary
# ---------------------------------------------------------------------------


def test_authorized_export_has_resumen_and_serie_diaria_sheets(db):
    """3b.1 (spec: "Authorized export includes both sheets") -- and the half
    that actually matters: a metric that is NOT complete appears by state and
    reason, with an EMPTY value cell. A zero would be read as a measurement of
    zero millimetres, which is the single most expensive lie a rainfall report
    can tell."""
    from app.domains.geo.rainfall.export import RESUMEN_SHEET, SERIE_SHEET
    from app.domains.geo.rainfall.service import metric_rows, normalize_snapshot

    revision = _seeded_revision(db, scope_id="zone-3b1-sheets")
    response = _client(db).get(f"/rainfall/analyses/{revision.id}.xlsx")

    assert response.status_code == 200
    book = _sheets(response.content)
    assert book.sheetnames == [RESUMEN_SHEET, SERIE_SHEET]

    normalized = normalize_snapshot(
        revision.snapshot, expected_policy_revision=revision.policy_revision
    )
    table = _metric_table(_rows(book[RESUMEN_SHEET]))
    # EVERY metric the CSV projection carries is present -- a suppressed one is
    # never dropped from the friendly view, which would hide the suppression
    # instead of disclosing it.
    assert len(table) == len(metric_rows(normalized))

    states = {row["Estado"] for row in table}
    assert len(states) > 1, table  # the fixture is meant to disclose a mix
    for row in table:
        if row["Estado"] != "disponible":
            assert row["Valor"] is None, row
            assert isinstance(row["Motivo"], str) and row["Motivo"], row

    series_rows = _rows(book[SERIE_SHEET])
    assert len(series_rows) > 1
    assert series_rows[1][0] == f"{_CURVE_YEAR}-01-01"


def test_unauthorized_export_denied(db):
    """3b.2 (spec: "Unauthorized export is denied") -- proven with NO
    dependency override, so the router-level `require_admin_or_operator` is
    what answers, and it answers BEFORE the handler: an unauthenticated caller
    cannot even learn whether the revision exists."""
    from uuid import uuid4

    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from app.domains.geo.rainfall.router import router

    revision = _seeded_revision(db, scope_id="zone-3b2-denied")
    app = FastAPI()
    app.include_router(router)
    client = TestClient(app)

    assert client.get(f"/rainfall/analyses/{revision.id}.xlsx").status_code == 401
    # Same answer for a revision that does not exist: no existence oracle.
    assert client.get(f"/rainfall/analyses/{uuid4()}.xlsx").status_code == 401


def test_unknown_revision_is_a_404_for_an_authorized_caller(db):
    """The CSV route's own 404 semantics, inherited rather than restated."""
    from uuid import uuid4

    assert _client(db).get(f"/rainfall/analyses/{uuid4()}.xlsx").status_code == 404


# ---------------------------------------------------------------------------
# 3b.3: the consistency stamp, both directions
# ---------------------------------------------------------------------------


def test_resumen_stamps_series_consistency_both_directions(db):
    """3b.3 (design.md D3/D7) -- the workbook outlives the screen that showed
    the staleness notice, so the Resumen sheet stamps the pin itself. Both
    directions are exercised on the SAME revision: consistent first, then the
    same revision after a slot inside its window is superseded."""
    from app.domains.geo.rainfall.export import CONSISTENCY_LABEL, RESUMEN_SHEET
    from app.domains.geo.rainfall.models import RainfallIntervalLifecycle

    revision = _seeded_revision(db, scope_id="zone-3b3-stamp")
    client = _client(db)

    before = _labelled(
        _rows(_sheets(client.get(f"/rainfall/analyses/{revision.id}.xlsx").content)[RESUMEN_SHEET])
    )
    assert before[CONSISTENCY_LABEL] == "sí"

    # An NRT correction supersedes one day inside the analysis window, exactly
    # as `test_superseded_slot_reports_data_revision_moved` plants it.
    _persist_zone_rows(
        db,
        scope_id="zone-3b3-stamp",
        rows=_daily_rows(date(_CURVE_YEAR, 1, 5), 1, 9.0, provider_revision=_ZONE_FAMILY),
    )
    # A real SUPERSESSION, not a second live family: the two produce different
    # `consistency_reason` values, so asserting the precondition is what makes
    # the stamp below a statement about `data_revision_moved` in particular.
    assert (
        db.query(RainfallIntervalLifecycle)
        .filter(RainfallIntervalLifecycle.superseded_by_id.isnot(None))
        .count()
        > 0
    )

    after = _labelled(
        _rows(_sheets(client.get(f"/rainfall/analyses/{revision.id}.xlsx").content)[RESUMEN_SHEET])
    )
    # design.md D7's literal contract string, reconstructed from the two cells:
    # "Serie diaria consistente con el análisis: sí | no — <motivo>".
    assert after[CONSISTENCY_LABEL] == "no — data_revision_moved"
    assert (
        f"{CONSISTENCY_LABEL}: {after[CONSISTENCY_LABEL]}"
        == "Serie diaria consistente con el análisis: no — data_revision_moved"
    )


# ---------------------------------------------------------------------------
# 3b.4: transport headers
# ---------------------------------------------------------------------------


def test_export_filename_and_content_disposition(db):
    """3b.4 -- the browser must save it as a file with a name that identifies
    the revision, and the media type must be the xlsx one or Excel refuses to
    open what it is handed."""
    from app.domains.geo.rainfall.export import XLSX_MEDIA_TYPE

    revision = _seeded_revision(db, scope_id="zone-3b4-headers")
    response = _client(db).get(f"/rainfall/analyses/{revision.id}.xlsx")

    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert (
        response.headers["content-disposition"]
        == f'attachment; filename="lluvia_{revision.id}.xlsx"'
    )
    # A real workbook, not an HTML error page with an optimistic header.
    assert response.content[:2] == b"PK"


# ---------------------------------------------------------------------------
# 3b.5: the audit CSV is untouched
# ---------------------------------------------------------------------------


def test_audit_csv_bytes_unchanged(db):
    """3b.5 (spec: "The existing audit CSV route and contract MUST remain
    unchanged") -- the xlsx is an ADDITIONAL friendly view, never a
    redefinition of the audit artifact. Regression-proven at the byte level,
    including that the CSV never acquired the xlsx route's download headers."""
    from app.domains.geo.rainfall.service import metric_rows, metric_rows_csv, normalize_snapshot

    revision = _seeded_revision(db, scope_id="zone-3b5-csv")
    client = _client(db)

    csv_response = client.get(f"/rainfall/analyses/{revision.id}.csv")
    client.get(f"/rainfall/analyses/{revision.id}.xlsx")
    csv_again = client.get(f"/rainfall/analyses/{revision.id}.csv")

    assert csv_response.status_code == 200
    assert csv_response.content == csv_again.content
    assert csv_response.headers["content-type"].startswith("text/csv")
    assert "content-disposition" not in csv_response.headers
    expected = metric_rows_csv(
        metric_rows(
            normalize_snapshot(revision.snapshot, expected_policy_revision=revision.policy_revision)
        )
    )
    assert csv_response.text == expected


# ---------------------------------------------------------------------------
# The consistency contract: the xlsx cannot silently diverge from the JSON
# ---------------------------------------------------------------------------


def test_xlsx_metric_values_match_the_served_csv_exactly(db):
    """The xlsx and the audit CSV are two renderings of ONE projection
    (``metric_rows(normalize_snapshot(...))``, design.md D7). Compared as two
    SERVED artifacts rather than against a hand-built expectation, because a
    hand-built expectation cannot catch the failure this guards: the friendly
    sheet recomputing, rounding or zero-filling a number the audit trail
    reports differently."""
    from app.domains.geo.rainfall.export import RESUMEN_SHEET

    revision = _seeded_revision(db, scope_id="zone-3b-parity")
    client = _client(db)

    csv_rows = list(
        csv.DictReader(StringIO(client.get(f"/rainfall/analyses/{revision.id}.csv").text))
    )
    table = _metric_table(
        _rows(_sheets(client.get(f"/rainfall/analyses/{revision.id}.xlsx").content)[RESUMEN_SHEET])
    )

    assert len(table) == len(csv_rows)
    for sheet_row, audit_row in zip(table, csv_rows, strict=True):
        audit_value = audit_row["value"]
        if audit_value == "":
            # Unknown stays unknown: an empty CELL, never a zero.
            assert sheet_row["Valor"] is None, sheet_row
        else:
            assert sheet_row["Valor"] == pytest.approx(float(audit_value)), sheet_row
        assert (sheet_row["Motivo"] or "") == audit_row["reason"]
        assert (sheet_row["Unidad"] or "") == audit_row["unit"]


def test_serie_diaria_matches_the_served_series_json_point_for_point(db):
    """Same rule for the second sheet: the "Serie diaria" rows and the
    ``/series`` JSON come from ONE ``build_series`` contract, so a day the
    chart shows as unknown cannot be a zero in the workbook the same operator
    downloads five seconds later."""
    from app.domains.geo.rainfall.export import SERIE_SHEET, series_table_header

    revision = _seeded_revision(db, scope_id="zone-3b-series-parity")
    client = _client(db)

    served = client.get(f"/rainfall/analyses/{revision.id}/series").json()
    sheet = _rows(
        _sheets(client.get(f"/rainfall/analyses/{revision.id}.xlsx").content)[SERIE_SHEET]
    )

    assert sheet[0] == list(series_table_header(served["unit"]))
    assert len(sheet) - 1 == len(served["points"])
    for row, point in zip(sheet[1:], served["points"], strict=True):
        assert row[0] == point["date"]
        for column, field in ((1, "mm"), (2, "accumulated"), (3, "normal_accumulated")):
            if point[field] is None:
                assert row[column] is None, (row, point)
            else:
                assert row[column] == pytest.approx(point[field]), (row, point)


def test_resumen_names_the_last_day_with_evidence_not_the_exclusive_end(db):
    """JDA-001 / JDB-001 -- "Datos disponibles hasta" is an INCLUSIVE claim
    and ``available_through`` is an EXCLUSIVE bound.

    ``compute._disclosure_window`` builds ``available_through`` as
    ``min(comparison_end + 1 day, max(interval_end))`` and ``series._points``
    stops one day short of it, so stamping the raw value under this label
    claims a day the workbook itself has no row for -- and on a finalized past
    year it names a day in the NEXT year, which the analysis says nothing
    about. The reader keeps this file, so the sentence has to survive without
    the screen that would have qualified it.

    Asserted against the workbook's OWN last Serie diaria row rather than
    against a recomputed date: the two cells must tell one story, and a test
    that re-derived the subtraction here would pass on any arithmetic the
    export happened to use.
    """
    from app.domains.geo.rainfall.export import (
        AVAILABLE_THROUGH_LABEL,
        RESUMEN_SHEET,
        SERIE_SHEET,
    )

    revision = _seeded_revision(db, scope_id="zone-3b-available-through")
    client = _client(db)
    content = client.get(f"/rainfall/analyses/{revision.id}.xlsx").content
    sheets = _sheets(content)

    block = _labelled(_rows(sheets[RESUMEN_SHEET]))
    serie = _rows(sheets[SERIE_SHEET])
    last_day_with_evidence = serie[-1][0]
    served = client.get(f"/rainfall/analyses/{revision.id}/series").json()

    assert block[AVAILABLE_THROUGH_LABEL] == last_day_with_evidence
    # …and it is genuinely the day BEFORE the wire value, not a coincidence of
    # this fixture: the exclusive bound must not appear in the cell at all.
    assert served["available_through"][:10] != last_day_with_evidence
    assert served["available_through"] not in str(block[AVAILABLE_THROUGH_LABEL])


def test_a_non_utc_available_through_still_names_the_same_last_day(db):
    """JDB-101 -- the LI3A-005 defect class, re-opened on the disclosure
    window itself.

    ``available_through`` is ``_disclosure_window``'s ``window_end``
    serialized, and under provider lag that value is ``max(interval_end)`` --
    a ``timestamptz`` ``psycopg2`` renders in the database session's own
    ``TimeZone``, which nothing in this repository pins. So the stored string
    can legitimately carry a non-UTC offset for the very same instant, and
    every consumer that reads a calendar day off its first ten characters
    lands a day early: ``2024-03-02T21:00:00-03:00`` IS
    ``2024-03-03T00:00:00+00:00``, but ``.date()`` on it says March 2, and
    minus one day says March 1.

    Both display consumers are covered by one instant here: the workbook cell
    (asserted against the workbook's OWN last Serie diaria row, which is
    driven by the same window and therefore does not move) and the ``/series``
    echo the chart reads, which must leave the backend already normalized.
    """
    from app.domains.geo.rainfall.export import (
        AVAILABLE_THROUGH_LABEL,
        RESUMEN_SHEET,
        SERIE_SHEET,
        build_workbook,
    )
    from app.domains.geo.rainfall.series import build_series

    revision = _seeded_revision(db, scope_id="zone-3b-available-through-tz")
    # Same INSTANT as the built envelope (2024-03-03T00:00+00:00), rendered in
    # UTC-3. Injected through the revision-shaped stand-in the slice-3a tests
    # use, because a stored row is append-only and ORM-guarded against update.
    shifted = {**revision.snapshot}
    annual = {**shifted["annual"]}
    selected = {**annual["selected"]}
    provenance = {**selected["provenance"]}
    assert provenance["available_through"] == datetime(2024, 3, 3, tzinfo=UTC).isoformat()
    provenance["available_through"] = "2024-03-02T21:00:00-03:00"
    selected["provenance"] = provenance
    annual["selected"] = selected
    shifted["annual"] = annual
    stand_in = _fake_revision(shifted, data_revision=revision.data_revision)
    # `build_workbook` also reads the policy revision the row was written
    # under; the stand-in carries the real one so nothing else moves.
    stand_in.policy_revision = revision.policy_revision

    series = build_series(db, stand_in)
    sheets = _sheets(build_workbook(db, stand_in).content)
    block = _labelled(_rows(sheets[RESUMEN_SHEET]))
    last_day_with_evidence = _rows(sheets[SERIE_SHEET])[-1][0]

    # The window did not move -- only its rendering did.
    assert last_day_with_evidence == "2024-03-02"
    assert block[AVAILABLE_THROUGH_LABEL] == last_day_with_evidence
    # The wire value leaves the backend in UTC, so the client's own day
    # arithmetic (a slice of the first ten characters) cannot inherit the
    # session's timezone from a string it has no way to normalize.
    assert series["available_through"] == datetime(2024, 3, 3, tzinfo=UTC).isoformat()


def test_a_snapshot_too_broken_to_export_is_refused_not_rendered(db):
    """Divergence, the other direction: when the stored envelope cannot back a
    series at all, the xlsx REFUSES with the same 503 the CSV and `/series`
    routes give, rather than shipping a half-built workbook. Silence would be
    the worst outcome here -- a downloaded file has no error banner."""
    from uuid import uuid4

    from app.domains.geo.rainfall.repository import RainfallRepository

    class _Broken:
        id = uuid4()
        data_revision = "d" * 64
        policy_revision = "rainfall-v2-2026-08-insights"
        snapshot = {"scope": {"kind": "zone", "id": "z", "version": "v1"}, "year": 2025}

    broken = _Broken()
    original = RainfallRepository.get_revision
    try:
        RainfallRepository.get_revision = lambda self, session, revision_id: broken
        response = _client(db).get(f"/rainfall/analyses/{broken.id}.xlsx")
    finally:
        RainfallRepository.get_revision = original

    assert response.status_code == 503


# ---------------------------------------------------------------------------
# The normal-curve state, disclosed rather than rendered as blank columns
# ---------------------------------------------------------------------------


def test_resumen_stamps_the_normal_curve_state(db):
    """LI3A-001 carried into the workbook: a REFUSED curve and a structurally
    ABSENT one both render every "Normal acumulada" cell empty, so without a
    status cell the file cannot tell its reader which of the two it is holding
    -- the exact ambiguity `normal_curve_state` was added to the wire contract
    to remove. The two labels must therefore differ."""
    from app.domains.geo.rainfall.export import (
        NORMAL_CURVE_LABEL,
        NORMAL_CURVE_STATE_LABELS,
        RESUMEN_SHEET,
        SERIE_SHEET,
    )
    from app.domains.geo.rainfall.models import RainfallIntervalValue

    revision = _seeded_revision(db, scope_id="zone-3b-curve-refused")
    client = _client(db)

    available = _labelled(
        _rows(_sheets(client.get(f"/rainfall/analyses/{revision.id}.xlsx").content)[RESUMEN_SHEET])
    )
    assert available[NORMAL_CURVE_LABEL] == NORMAL_CURVE_STATE_LABELS["available"]

    # A duplicated baseline slot lands AFTER the build: the curve is computed
    # and then refused (the sibling of `baseline_cumulatives`' own guard).
    duplicated_day = datetime(1991, 1, 5, tzinfo=UTC)
    db.add(
        RainfallIntervalValue(
            source_id="chirps-v3-final",
            scope_kind="provider_asset",
            scope_id=asset_name_for("zone", "zone-3b-curve-refused"),
            scope_version=BASELINE_ASSET_VERSION,
            interval_start=duplicated_day,
            interval_end=duplicated_day + timedelta(days=1),
            provider_revision="v3-final+r1",
            value=500.0,
            unit="mm",
        )
    )
    db.flush()

    book = _sheets(client.get(f"/rainfall/analyses/{revision.id}.xlsx").content)
    refused = _labelled(_rows(book[RESUMEN_SHEET]))

    assert refused[NORMAL_CURVE_LABEL] == NORMAL_CURVE_STATE_LABELS["integrity_refused"]
    assert NORMAL_CURVE_STATE_LABELS["integrity_refused"] != NORMAL_CURVE_STATE_LABELS["suppressed"]
    # ... and the columns it governs really are blank, so the stamp is the ONLY
    # thing standing between the reader and an unexplained empty column.
    assert all(row[3] is None for row in _rows(book[SERIE_SHEET])[1:])


def test_resumen_stamps_a_suppressed_normal_curve_distinctly(db):
    """The honest-absence half of the pair above: no baseline persisted at all
    means `annual.normal` suppresses, and the workbook says so in its own
    words rather than reusing the refusal wording."""
    from app.domains.geo.rainfall.export import (
        NORMAL_CURVE_LABEL,
        NORMAL_CURVE_STATE_LABELS,
        RESUMEN_SHEET,
    )

    revision = _seeded_revision(db, scope_id="zone-3b-curve-suppressed", baseline=False)
    assert revision.snapshot["annual"]["normal"]["state"] == "suppressed"

    block = _labelled(
        _rows(
            _sheets(_client(db).get(f"/rainfall/analyses/{revision.id}.xlsx").content)[
                RESUMEN_SHEET
            ]
        )
    )

    assert block[NORMAL_CURVE_LABEL] == NORMAL_CURVE_STATE_LABELS["suppressed"]


def test_resumen_carries_the_disclosure_time_summary(db):
    """The summary is assembled at disclosure time from post-policy states
    (design.md D4) and the xlsx Resumen sheet is one of the two surfaces that
    carry it (the CSV's row projection cannot -- it never iterates root keys,
    LIA-102). Read from the same `normalize_snapshot` result, so it can never
    describe states this file does not show."""
    from app.domains.geo.rainfall.export import RESUMEN_SHEET, SUMMARY_LABEL
    from app.domains.geo.rainfall.service import normalize_snapshot

    revision = _seeded_revision(db, scope_id="zone-3b-summary")
    block = _labelled(
        _rows(
            _sheets(_client(db).get(f"/rainfall/analyses/{revision.id}.xlsx").content)[
                RESUMEN_SHEET
            ]
        )
    )

    expected = normalize_snapshot(
        revision.snapshot, expected_policy_revision=revision.policy_revision
    )["summary"]
    assert block[SUMMARY_LABEL] == expected


# ---------------------------------------------------------------------------
# Counterexample self-check: hostile text, a metric the policy reduced, and an
# analysis with no evidence at all
# ---------------------------------------------------------------------------


def _revision_stand_in(snapshot: dict, *, revision):
    """A revision-shaped stand-in carrying a doctored snapshot.

    Revision rows are append-only and ORM-guarded against update, so an
    envelope in a state the build would not produce cannot be created by
    writing one -- the same reason slice 3a's `_fake_revision` exists.
    """
    from types import SimpleNamespace

    return SimpleNamespace(
        id=revision.id,
        data_revision=revision.data_revision,
        policy_revision=revision.policy_revision,
        snapshot=snapshot,
    )


def _cells(sheet) -> list[list]:
    return [list(row) for row in sheet.iter_rows()]


def test_text_that_looks_like_a_formula_is_written_as_text(db):
    """Security: openpyxl types a string starting with ``=`` as a FORMULA
    (verified: `data_type == "f"`), which is the spreadsheet-injection class --
    a value that reaches a cell becomes executable content in the reader's
    Excel or LibreOffice, on a machine the server never sees.

    Every string this workbook writes is server-built TODAY (policy enum
    reasons, a Spanish narrative, ISO dates), so the vector is latent rather
    than open. It is guarded anyway, at the presentation layer, because the
    fields most likely to change are exactly the ones fed from provider batches
    (``discrepancies``) and from zoning configuration (``scope.id``) -- and a
    guard that depends on today's data being benign is not a guard.

    TWO LAYERS since LI3B-001. The value-level one is the sanitizer the audit
    CSV route shares (`service.neutralize_spreadsheet_formula`), so one hostile
    value renders identically in both exports of a revision; the structural
    ``data_type = "s"`` stays underneath it. They are asserted separately below
    because they fail independently: dropping ``WriteOnlyCell`` loses the
    second, and a prefix the sanitizer has not learned loses the first.
    """
    from app.domains.geo.rainfall.export import RESUMEN_SHEET, build_workbook

    revision = _seeded_revision(db, scope_id="zone-3b-formula")
    hostile = {**revision.snapshot}
    annual = {**hostile["annual"]}
    selected = {**annual["selected"]}
    selected["unit"] = "=1+1"
    selected["discrepancies"] = ['=HYPERLINK("http://evil.example/x","ver")']
    annual["selected"] = selected
    hostile["annual"] = annual

    workbook = build_workbook(db, _revision_stand_in(hostile, revision=revision))
    # Loaded WITHOUT `data_only`, because that is where the two halves of the
    # defect show up: an unguarded cell is typed `f`, and a data-consuming
    # reader (`data_only=True`, an uncached formula) then sees `None` where the
    # server stored a value -- injection AND silent data loss in one cell.
    book = load_workbook(BytesIO(workbook.content), read_only=True)
    rows = _cells(book[RESUMEN_SHEET])
    all_cells = [(cell.value, cell.data_type) for row in rows for cell in row]

    # Layer 1: neither hostile string reaches the file with its trigger
    # character in the leading position a reader evaluates.
    assert not [
        value for value, _type in all_cells if isinstance(value, str) and value.startswith("=")
    ], all_cells

    # Both are still THERE, whole, carrying the standard text marker -- an
    # export is evidence, so the value is prefixed, never stripped or rewritten.
    hostile_cells = [
        cell
        for row in rows
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("'=")
    ]
    assert len(hostile_cells) == 2, all_cells
    assert {cell.value for cell in hostile_cells} == {
        "'=1+1",
        '\'=HYPERLINK("http://evil.example/x","ver")',
    }
    # Layer 2: inline text, NOT a formula. Independent of layer 1 -- this is
    # what still holds if a future trigger character slips past the sanitizer.
    for cell in hostile_cells:
        assert cell.data_type == "s", (cell.value, cell.data_type)


def test_a_metric_the_policy_reduced_still_gets_a_row_with_empty_columns(db):
    """Absence: a metric that fails its contract at disclosure is rewritten by
    ``service._unavailable`` into a FOUR-field dict, losing unit, interval,
    coverage, provenance and the rest. The friendly sheet must still give it a
    row -- dropping it would hide the suppression instead of disclosing it --
    with the lost columns EMPTY rather than zero-filled."""
    from app.domains.geo.rainfall.export import RESUMEN_SHEET, build_workbook
    from app.domains.geo.rainfall.service import metric_rows, normalize_snapshot

    revision = _seeded_revision(db, scope_id="zone-3b-reduced")
    reduced = {**revision.snapshot}
    # An ANTECEDENT is doctored rather than `annual.selected`: the series
    # module reads the latter's provenance and would refuse the whole export
    # (proven by `test_a_snapshot_too_broken_to_export_is_refused_not_rendered`),
    # which is a different contract from the presentation one under test here.
    antecedents = {**reduced["antecedents"]}
    # A stored metric with no coverage/completeness: `_normalize_metric` cannot
    # validate it and rewrites it as `metric_contract_invalid`.
    antecedents["d7"] = {"metric": "d7", "value": 12.0, "state": "available"}
    reduced["antecedents"] = antecedents
    normalized = normalize_snapshot(reduced, expected_policy_revision=revision.policy_revision)
    assert normalized["antecedents"]["d7"] == {
        "metric": "d7",
        "value": None,
        "state": "unavailable",
        "reason": "metric_contract_invalid",
    }

    workbook = build_workbook(db, _revision_stand_in(reduced, revision=revision))
    table = _metric_table(_rows(_sheets(workbook.content)[RESUMEN_SHEET]))

    assert len(table) == len(metric_rows(normalized))
    row = next(row for row in table if row["Motivo"] == "metric_contract_invalid")
    assert row["Estado"] == "no disponible"
    assert row["Valor"] is None
    for column in ("Unidad", "Desde", "Hasta", "Cobertura", "Completitud", "Fuente"):
        assert row[column] is None, (column, row)


def test_an_analysis_with_no_evidence_still_exports_both_sheets(db):
    """Boundary: a revision built with NO interval evidence at all is a real,
    servable revision (`annual.selected.state: "unavailable"`). Its export must
    still be a valid workbook that discloses the emptiness -- an operator who
    downloads it has to be able to see that there is nothing there, which is
    not the same experience as a download that fails."""
    from app.domains.geo.rainfall.export import CONSISTENCY_LABEL, RESUMEN_SHEET, SERIE_SHEET

    now = datetime(2025, 1, 21, 12, 0, tzinfo=UTC)
    revision = _build_revision(db, scope_id="zone-3b-no-evidence", year=2025, now=now)
    assert revision.snapshot["annual"]["selected"]["state"] == "unavailable"

    book = _sheets(_client(db).get(f"/rainfall/analyses/{revision.id}.xlsx").content)

    # The zero-resolved-rows shape the observability workbook documents as
    # benign: permanently inconsistent, because there is no family to compare.
    assert (
        _labelled(_rows(book[RESUMEN_SHEET]))[CONSISTENCY_LABEL] == "no — interval_family_ambiguous"
    )
    series_rows = _rows(book[SERIE_SHEET])
    assert len(series_rows) == 22  # header + Jan 1..21, every day disclosed
    assert all(row[1] is None for row in series_rows[1:])


def test_xlsx_served_event_is_documented_in_the_observability_workbook():
    """LI2B-005's rule applied to this slice's own new event: an event that
    fires in production and appears nowhere in the catalogue is undocumented
    by construction."""
    from pathlib import Path

    workbook = (
        Path(__file__).resolve().parents[5] / "docs" / "lluvia-v2-observability-workbook.md"
    ).read_text(encoding="utf-8")

    assert "`rainfall.xlsx.served`" in workbook
    assert "normal_curve_state" in workbook
