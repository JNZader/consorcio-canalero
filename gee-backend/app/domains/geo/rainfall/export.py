"""The friendly xlsx view of one stored Rainfall v2 analysis (design.md D7).

Two sheets, two consumers of work that already exists:

* **Resumen** — the SAME ``metric_rows(normalize_snapshot(...))`` projection the
  audit CSV serializes, rendered with Spanish labels, plus the disclosures a
  downloaded file needs because it outlives the screen that showed them.
* **Serie diaria** — the slice-3a series builder (``series.build_series``), the
  same one ``GET /rainfall/analyses/{id}/series`` serves the chart.

Nothing here recomputes a number. That is the whole design: a workbook gets
forwarded, printed and argued over months after the tab that produced it was
closed, so a friendly view that silently disagrees with the audit trail is
worse than no friendly view at all. Parity is therefore STRUCTURAL — one
projection, one series contract, two renderings — rather than a pair of
implementations kept in step by review.

Three things a downloaded file must carry that a live screen gets for free:

1. ``None`` is written as an EMPTY CELL, never ``0``. A zero in a rainfall
   report reads as a measurement of zero millimetres; unknown must stay
   visibly unknown, exactly as the CSV leaves it blank.
2. The **series consistency pin** (design.md D3) is stamped in Resumen. The
   chart shows an `Alert` when ``consistent_with_snapshot`` is false; a
   workbook has no place to put a transient notice, so it states it as a row.
3. The **normal-curve state** is stamped too. ``suppressed`` and
   ``integrity_refused`` both render every "Normal acumulada" cell empty
   (LI3A-001), so without this row the reader cannot tell an analysis that has
   no baseline to compare against from one whose baseline was computed and
   discarded for contradicting the card beside it.

Boundary rule (design.md "Technical Approach"): ``repository.py`` owns SQL,
``compute.py`` stays pure, ``series.py`` owns the read-only Session for the
series, and this module owns presentation only. It is READ-ONLY: it writes
nothing and enqueues nothing.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, NamedTuple

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from sqlalchemy.orm import Session

from app.domains.geo.rainfall.models import RainfallAnalysisRevision
from app.domains.geo.rainfall.series import (
    NORMAL_CURVE_AVAILABLE,
    NORMAL_CURVE_INTEGRITY_REFUSED,
    NORMAL_CURVE_SUPPRESSED,
    build_series,
)
from app.domains.geo.rainfall.service import (
    SUMMARY_METRIC_LABELS,
    SUMMARY_STATE_LABELS,
    metric_rows,
    neutralize_spreadsheet_formula,
    normalize_snapshot,
)

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

RESUMEN_SHEET = "Resumen"
SERIE_SHEET = "Serie diaria"

TITLE = "Análisis de lluvia — Consorcio Canalero"

REVISION_LABEL = "Revisión del análisis"
SCOPE_LABEL = "Ámbito"
REGIONAL_LABEL = "Estimación regional"
YEAR_LABEL = "Año analizado"
COMPARISON_END_LABEL = "Fecha de comparación"
# An INCLUSIVE claim, so it is stamped with an inclusive day -- see
# `_last_evidence_day`, which is why the raw wire value never reaches this cell.
AVAILABLE_THROUGH_LABEL = "Datos disponibles hasta"
BASELINE_LABEL = "Período de referencia"
DATA_REVISION_LABEL = "Revisión de datos (data_revision)"
POLICY_REVISION_LABEL = "Revisión de política de métricas"
SUMMARY_LABEL = "Resumen del análisis"

# design.md D7's literal contract string is this label joined to its value:
# "Serie diaria consistente con el análisis: sí | no — <motivo>". Label and
# value live in two cells so the sheet stays a readable label/value grid; the
# sentence is the concatenation.
CONSISTENCY_LABEL = "Serie diaria consistente con el análisis"
CONSISTENCY_CONSISTENT = "sí"

NORMAL_CURVE_LABEL = "Curva normal 1991–2020"

# The three wire states, in the reader's language. They MUST read differently
# from one another (LI3A-001): the columns they describe are byte-identical in
# all three cases, so this cell is the only thing that distinguishes "there is
# no baseline for this analysis" from "there was one and it was discarded".
NORMAL_CURVE_STATE_LABELS: dict[str, str] = {
    NORMAL_CURVE_AVAILABLE: "disponible",
    NORMAL_CURVE_SUPPRESSED: "no disponible (este análisis no divulga una normal)",
    NORMAL_CURVE_INTEGRITY_REFUSED: (
        "descartada por una inconsistencia de datos (ver rainfall.series.normal_curve_refused)"
    ),
}

METRIC_TABLE_HEADER: tuple[str, ...] = (
    "Métrica",
    "Valor",
    "Unidad",
    "Estado",
    "Motivo",
    "Desde",
    "Hasta",
    "Cobertura",
    "Completitud",
    "Fuente",
    "Estado temporal",
    "Observaciones",
)

# `metric_rows` flattens the groups away, so the label is keyed on the wire
# `metric` name rather than on the group key `rainfall_summary` uses. Layered
# on top of that vocabulary rather than beside it: the same metric must be
# named the same way in the narrative, on the badge and in this sheet, which is
# the coherence rule the summary requirement states (spec delta, D4).
EXPORT_METRIC_LABELS: dict[str, str] = {
    **SUMMARY_METRIC_LABELS,
    "annual": SUMMARY_METRIC_LABELS["selected"],
    "annual_normal": SUMMARY_METRIC_LABELS["normal"],
    "annual_percentile": SUMMARY_METRIC_LABELS["percentile"],
}


def series_table_header(unit: str) -> tuple[str, ...]:
    """The Serie diaria header, carrying the analysis' OWN unit rather than a
    hardcoded "mm" that would quietly lie if the unit ever changed."""
    return (
        "Fecha",
        f"Lluvia diaria ({unit})",
        f"Acumulado ({unit})",
        f"Normal acumulada ({unit})",
        "Estado",
    )


def _append(sheet: Any, values: list[Any]) -> None:
    """Append one row, forcing every STRING cell to be inline text.

    *sheet* is a write-only worksheet. It is annotated ``Any`` on purpose: its
    class lives in ``openpyxl.worksheet._write_only``, a private module, and
    pinning this file to a private import path would be a worse coupling than
    the missing annotation.

    openpyxl types a string starting with ``=`` as a FORMULA (measured:
    ``data_type == "f"``), which is the spreadsheet-injection class -- the cell
    becomes executable content in the reader's Excel or LibreOffice, on a
    machine this server never sees. It is also silent data loss in the same
    breath: a data-consuming reader (``data_only=True``) finds no cached result
    and reads ``None`` where the server stored a value.

    Every string written here is server-built today (policy enum reasons, a
    Spanish narrative, ISO dates), so the vector is latent rather than open.
    Guarded regardless, and at the ONE place every cell passes through, because
    the fields most likely to change are precisely the ones fed from provider
    batches (``discrepancies``) and zoning configuration (``scope.id``) -- and
    a guard that depends on today's data being benign is not a guard.

    Numbers, ``None`` and dates are passed through untouched: forcing them to
    text would make the workbook unusable for the arithmetic it exists for.

    TWO LAYERS, and they are lost independently (LI3B-001). The value-level
    guard is ``service.neutralize_spreadsheet_formula`` -- the SAME sanitizer
    the audit CSV route uses, so the same hostile value renders identically in
    both exports of one revision instead of two files of the same analysis
    disagreeing. ``data_type = "s"`` stays as the second layer: it is what makes
    the guarantee structural here, and it survives a value the sanitizer's
    prefix set has not learned about yet. Conversely the sanitizer survives a
    refactor away from ``WriteOnlyCell``, which is the only thing holding the
    structural layer up.
    """
    row: list[Any] = []
    for value in values:
        if isinstance(value, str):
            cell = WriteOnlyCell(sheet, value=neutralize_spreadsheet_formula(value))
            cell.data_type = "s"
            row.append(cell)
        else:
            row.append(value)
    sheet.append(row)


def _label(mapping: dict[str, str], key: object) -> str:
    """Never blank: an unmapped key falls through to its own wire value, the
    same rule `metricLabel` follows on the client. A missing translation must
    degrade to an untranslated fact, never to an empty cell that reads as
    "nothing to report"."""
    return mapping.get(key, str(key)) if isinstance(key, str) else str(key)


def _last_evidence_day(available_through: str) -> str:
    """The INCLUSIVE last day the provider published, from the EXCLUSIVE bound.

    JDA-001 / JDB-001. ``series["available_through"]`` is the exclusive end of
    the disclosure window -- ``compute._disclosure_window`` builds it as
    ``min(comparison_end + 1 day, max(interval_end))`` and ``series._points``
    stops one day short of it, so the Serie diaria sheet in this same workbook
    has no row for that day. Stamping it under an inclusive label ("Datos
    disponibles hasta") therefore claims evidence the file itself contradicts,
    and on a finalized past year it names January 1 of the FOLLOWING year,
    which the analysis says nothing about.

    A workbook outlives the screen that could have qualified it, so the cell
    has to be right on its own. The chart's footer converts the same value the
    same way, so the two surfaces keep telling one story.

    Parsing cannot fail here: ``build_series`` already refused a snapshot whose
    ``available_through`` is unparseable (``series._analysis`` ->
    ``SnapshotContractError`` -> 503), so this only ever sees a value that
    round-tripped through ``datetime.fromisoformat`` upstream.
    """
    return (datetime.fromisoformat(available_through).date() - timedelta(days=1)).isoformat()


def _consistency_value(consistent: bool, reason: str | None) -> str:
    if consistent:
        return CONSISTENCY_CONSISTENT
    # `consistency_reason` is non-null whenever the flag is false, by
    # construction (series.py); the fallback exists so a future fourth state
    # cannot produce a bare "no" with no explanation.
    return f"no — {reason}" if reason else "no"


def _metric_sheet_row(metric: dict[str, Any]) -> list[Any]:
    """One flattened metric, as the reader sees it.

    Every lookup is a ``.get``: a metric the policy downgraded is rewritten by
    ``service._unavailable`` into a four-field dict (metric/value/state/reason),
    so most columns are legitimately absent for it — and it still gets a ROW,
    because dropping it would hide the suppression instead of disclosing it.
    """
    provenance = metric.get("provenance")
    discrepancies = metric.get("discrepancies")
    return [
        _label(EXPORT_METRIC_LABELS, metric.get("metric")),
        metric.get("value"),
        metric.get("unit"),
        _label(SUMMARY_STATE_LABELS, metric.get("state")),
        metric.get("reason"),
        metric.get("interval_start"),
        metric.get("interval_end"),
        metric.get("coverage"),
        metric.get("completeness"),
        provenance.get("source_id") if isinstance(provenance, dict) else None,
        metric.get("temporal_state"),
        "; ".join(str(item) for item in discrepancies) if discrepancies else None,
    ]


class RainfallWorkbook(NamedTuple):
    """The bytes, plus what the route needs to describe what it just served
    without re-opening the file it built."""

    content: bytes
    consistent_with_snapshot: bool
    consistency_reason: str | None
    normal_curve_state: str
    points: int


def build_workbook(db: Session, revision: RainfallAnalysisRevision) -> RainfallWorkbook:
    """Render one stored revision as a two-sheet workbook.

    Raises :class:`service.SnapshotContractError` — from ``normalize_snapshot``
    or from ``build_series`` — for a stored envelope that cannot describe
    itself, which the route maps to the same 503 the CSV and ``/series`` routes
    already give. Refusing is the only safe answer here: a downloaded file has
    no error banner, so a half-built workbook would be read as a complete one.
    """
    snapshot = revision.snapshot
    normalized = normalize_snapshot(snapshot, expected_policy_revision=revision.policy_revision)
    series = build_series(db, revision)

    workbook = Workbook(write_only=True)
    resumen = workbook.create_sheet(title=RESUMEN_SHEET)

    scope = snapshot.get("scope") if isinstance(snapshot, dict) else None
    scope_text = (
        f"{scope.get('kind')} · {scope.get('id')} · {scope.get('version')}"
        if isinstance(scope, dict)
        else None
    )
    _append(resumen, [TITLE])
    for label, value in (
        (REVISION_LABEL, str(revision.id)),
        (SCOPE_LABEL, scope_text),
        (REGIONAL_LABEL, "sí" if snapshot.get("regional_estimate") else "no"),
        (YEAR_LABEL, series["year"]),
        (COMPARISON_END_LABEL, series["comparison_end"]),
        (AVAILABLE_THROUGH_LABEL, _last_evidence_day(series["available_through"])),
        (BASELINE_LABEL, snapshot.get("baseline")),
        (DATA_REVISION_LABEL, series["data_revision"]),
        (POLICY_REVISION_LABEL, revision.policy_revision),
        (SUMMARY_LABEL, normalized.get("summary")),
        (
            CONSISTENCY_LABEL,
            _consistency_value(series["consistent_with_snapshot"], series["consistency_reason"]),
        ),
        (NORMAL_CURVE_LABEL, _label(NORMAL_CURVE_STATE_LABELS, series["normal_curve_state"])),
    ):
        _append(resumen, [label, value])
    _append(resumen, [])
    _append(resumen, list(METRIC_TABLE_HEADER))
    for metric in metric_rows(normalized):
        _append(resumen, _metric_sheet_row(metric))

    serie = workbook.create_sheet(title=SERIE_SHEET)
    _append(serie, list(series_table_header(series["unit"])))
    for point in series["points"]:
        _append(
            serie,
            [
                point["date"],
                point["mm"],
                point["accumulated"],
                point["normal_accumulated"],
                _label(SUMMARY_STATE_LABELS, point["state"]),
            ],
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return RainfallWorkbook(
        content=buffer.getvalue(),
        consistent_with_snapshot=series["consistent_with_snapshot"],
        consistency_reason=series["consistency_reason"],
        normal_curve_state=series["normal_curve_state"],
        points=len(series["points"]),
    )
