"""Business-logic layer for padron domain."""

import csv
import io
import re
import uuid
from typing import Any, Optional

from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.domains.padron.models import Consorcista
from app.domains.padron.repository import PadronRepository
from app.domains.padron.schemas import ConsorcistaCreate, ConsorcistaUpdate


_NON_DIGITS = re.compile(r"\D")

# Column aliases for CSV/XLSX import (same logic as old padron_service.py)
_COLUMN_ALIASES: dict[str, set[str]] = {
    "nombre": {"nombre", "nombres", "name"},
    "apellido": {"apellido", "apellidos", "surname"},
    "cuit": {"cuit", "cuil", "dni", "documento", "doc"},
    "email": {"email", "correo", "mail"},
    "telefono": {"telefono", "tel", "celular", "movil"},
    "domicilio": {"direccion", "direccion_postal", "domicilio"},
    "localidad": {"localidad", "ciudad", "city"},
    "parcela": {"parcela", "lote"},
    "hectareas": {"hectareas", "has", "superficie"},
    "categoria": {"categoria", "tipo"},
}


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_cuit(value: Any) -> Optional[str]:
    raw = _normalize_text(value)
    if not raw:
        return None
    digits = _NON_DIGITS.sub("", raw)
    if len(digits) == 11:
        return f"{digits[:2]}-{digits[2:10]}-{digits[10:]}"
    return raw


def _canonical_key(raw_key: Any) -> Optional[str]:
    normalized = _normalize_header(raw_key)
    for canonical, aliases in _COLUMN_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


class PadronService:
    """Orchestrates repository calls with business rules."""

    def __init__(self, repository: PadronRepository | None = None) -> None:
        self.repo = repository or PadronRepository()

    # ── QUERIES ───────────────────────────────

    def get_by_id(self, db: Session, consorcista_id: uuid.UUID) -> Consorcista:
        consorcista = self.repo.get_by_id(db, consorcista_id)
        if consorcista is None:
            raise HTTPException(status_code=404, detail="Consorcista no encontrado")
        return consorcista

    def list_consorcistas(
        self,
        db: Session,
        *,
        page: int = 1,
        limit: int = 20,
        estado: Optional[str] = None,
        categoria: Optional[str] = None,
        search: Optional[str] = None,
    ) -> tuple[list[Consorcista], int]:
        return self.repo.get_all(
            db,
            page=page,
            limit=limit,
            estado_filter=estado,
            categoria_filter=categoria,
            search=search,
        )

    def get_stats(self, db: Session) -> dict:
        return self.repo.get_stats(db)

    # ── COMMANDS ──────────────────────────────

    def create(self, db: Session, data: ConsorcistaCreate) -> Consorcista:
        """Create a consorcista, checking for duplicate CUIT."""
        existing = self.repo.get_by_cuit(db, data.cuit)
        if existing is not None:
            raise HTTPException(
                status_code=409,
                detail=f"Ya existe un consorcista con CUIT {data.cuit}",
            )
        consorcista = self.repo.create(db, data)
        db.commit()
        db.refresh(consorcista)
        return consorcista

    def update(
        self,
        db: Session,
        consorcista_id: uuid.UUID,
        data: ConsorcistaUpdate,
    ) -> Consorcista:
        """Update an existing consorcista, checking CUIT uniqueness if changed."""
        self.get_by_id(db, consorcista_id)

        # Check CUIT uniqueness if being updated
        update_fields = data.model_dump(exclude_unset=True)
        if "cuit" in update_fields and update_fields["cuit"] is not None:
            existing = self.repo.get_by_cuit(db, update_fields["cuit"])
            if existing is not None and existing.id != consorcista_id:
                raise HTTPException(
                    status_code=409,
                    detail=f"Ya existe un consorcista con CUIT {update_fields['cuit']}",
                )

        updated = self.repo.update(db, consorcista_id, data)
        db.commit()
        db.refresh(updated)  # type: ignore[arg-type]
        return updated  # type: ignore[return-value]

    # ── CSV/XLSX IMPORT ──────────────────────

    def import_csv(
        self,
        db: Session,
        file_content: bytes,
        filename: str,
    ) -> dict[str, Any]:
        """
        Parse CSV/XLSX, validate CUITs, and bulk create consorcistas.
        Skips rows with duplicate CUITs (already in DB or within the file).
        """
        rows = self._parse_rows(filename, file_content)

        processed = 0
        created = 0
        skipped = 0
        errors: list[dict[str, Any]] = []
        seen_cuits: set[str] = set()

        for row_number, row in rows:
            processed += 1
            try:
                payload = self._build_payload(row)

                if not payload.get("nombre") or not payload.get("apellido"):
                    skipped += 1
                    errors.append({
                        "row": row_number,
                        "error": "Nombre y apellido son obligatorios",
                    })
                    continue

                cuit = payload.get("cuit")
                if not cuit:
                    skipped += 1
                    errors.append({
                        "row": row_number,
                        "error": "CUIT/CUIL es obligatorio",
                    })
                    continue

                # Duplicate within file
                if cuit in seen_cuits:
                    skipped += 1
                    errors.append({
                        "row": row_number,
                        "error": f"CUIT {cuit} duplicado en el archivo",
                    })
                    continue

                # Duplicate in DB
                if self.repo.get_by_cuit(db, cuit) is not None:
                    skipped += 1
                    errors.append({
                        "row": row_number,
                        "error": f"CUIT {cuit} ya existe en el padron",
                    })
                    continue

                seen_cuits.add(cuit)

                schema = ConsorcistaCreate(
                    nombre=payload["nombre"],
                    apellido=payload["apellido"],
                    cuit=cuit,
                    domicilio=payload.get("domicilio"),
                    localidad=payload.get("localidad"),
                    telefono=payload.get("telefono"),
                    email=payload.get("email"),
                    parcela=payload.get("parcela"),
                    hectareas=(
                        float(payload["hectareas"])
                        if payload.get("hectareas")
                        else None
                    ),
                    categoria=payload.get("categoria"),
                )
                self.repo.create(db, schema)
                created += 1
            except Exception as exc:
                skipped += 1
                errors.append({"row": row_number, "error": str(exc)})

        if created > 0:
            db.commit()

        return {
            "processed": processed,
            "created": created,
            "skipped": skipped,
            "errors": errors,
        }

    def _build_payload(self, row: dict[str, Any]) -> dict[str, Any]:
        """Map raw CSV row keys to canonical field names."""
        mapped: dict[str, Any] = {}
        for key, value in row.items():
            canonical = _canonical_key(key)
            if not canonical:
                continue
            mapped[canonical] = value

        return {
            "nombre": _normalize_text(mapped.get("nombre")),
            "apellido": _normalize_text(mapped.get("apellido")),
            "cuit": _normalize_cuit(mapped.get("cuit")),
            "domicilio": _normalize_text(mapped.get("domicilio")),
            "localidad": _normalize_text(mapped.get("localidad")),
            "telefono": _normalize_text(mapped.get("telefono")),
            "email": _normalize_text(mapped.get("email")),
            "parcela": _normalize_text(mapped.get("parcela")),
            "hectareas": _normalize_text(mapped.get("hectareas")),
            "categoria": _normalize_text(mapped.get("categoria")),
        }

    def _parse_rows(
        self, filename: str, content: bytes
    ) -> list[tuple[int, dict[str, Any]]]:
        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            return self._parse_csv(content)
        if lower_name.endswith(".xlsx"):
            return self._parse_xlsx(content)
        if lower_name.endswith(".xls"):
            return self._parse_xls(content)
        raise ValueError("Formato no soportado. Usar CSV, XLS o XLSX")

    def _parse_csv(self, content: bytes) -> list[tuple[int, dict[str, Any]]]:
        decoded = None
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded is None:
            raise ValueError("No se pudo decodificar CSV. Use UTF-8 o Latin-1")

        reader = csv.DictReader(io.StringIO(decoded))
        rows: list[tuple[int, dict[str, Any]]] = []
        for index, row in enumerate(reader, start=2):
            rows.append((index, dict(row)))
        return rows

    def _parse_xlsx(self, content: bytes) -> list[tuple[int, dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise ValueError("Dependencia faltante para XLSX: openpyxl") from exc

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []

        headers_list = [str(value or "").strip() for value in headers]
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_number, values in enumerate(rows_iter, start=2):
            row_dict = {
                headers_list[idx]: values[idx] if idx < len(values) else None
                for idx in range(len(headers_list))
            }
            rows.append((row_number, row_dict))
        return rows

    def _parse_xls(self, content: bytes) -> list[tuple[int, dict[str, Any]]]:
        try:
            import xlrd
        except ImportError as exc:
            raise ValueError("Dependencia faltante para XLS: xlrd") from exc

        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return []

        headers = [
            str(sheet.cell_value(0, col) or "").strip()
            for col in range(sheet.ncols)
        ]
        rows: list[tuple[int, dict[str, Any]]] = []
        for row_idx in range(1, sheet.nrows):
            values = [
                sheet.cell_value(row_idx, col) for col in range(sheet.ncols)
            ]
            row_dict = {headers[idx]: value for idx, value in enumerate(values)}
            rows.append((row_idx + 1, row_dict))
        return rows
