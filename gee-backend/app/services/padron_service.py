"""
Padron Service.
Manages consortium members and their annual fee payments.
"""

import csv
import io
import re
from functools import lru_cache
from typing import List, Dict, Any, Optional, Iterable, Tuple
from uuid import UUID

from app.services.supabase_service import get_supabase_service
from app.core.logging import get_logger

logger = get_logger(__name__)

# Characters that have special meaning in PostgREST filter strings
_POSTGREST_SPECIAL_CHARS = re.compile(r"[,.()\[\]{}]")
_CUIT_DIGITS = re.compile(r"\D")

_COLUMN_ALIASES = {
    "nombre": {"nombre", "nombres", "name"},
    "apellido": {"apellido", "apellidos", "surname"},
    "cuit": {"cuit", "cuil", "dni", "documento", "doc"},
    "representa_a": {"representa_a", "representa", "empresa", "establecimiento"},
    "email": {"email", "correo", "mail"},
    "telefono": {"telefono", "tel", "celular", "movil"},
    "direccion_postal": {"direccion", "direccion_postal", "domicilio"},
    "activo": {"activo", "estado", "habilitado"},
}


def _sanitize_search(value: str) -> str:
    """
    Strip whitespace and remove characters that could break PostgREST
    filter syntax (commas, dots, parentheses, brackets, braces).

    Args:
        value: Raw search string from the user.

    Returns:
        Sanitized string safe for interpolation into an or_() filter.
    """
    cleaned = value.strip()
    cleaned = _POSTGREST_SPECIAL_CHARS.sub("", cleaned)
    return cleaned


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "_")


def _normalize_text(value: Any) -> Optional[str]:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_cuit(value: Any) -> Optional[str]:
    raw = _normalize_text(value)
    if not raw:
        return None
    digits = _CUIT_DIGITS.sub("", raw)
    if len(digits) == 11:
        return f"{digits[:2]}-{digits[2:10]}-{digits[10:]}"
    return raw


def _normalize_active(value: Any) -> bool:
    if value is None:
        return True
    text = str(value).strip().lower()
    if text in {"0", "false", "f", "no", "n", "inactivo", "baja"}:
        return False
    return True


def _canonical_key(raw_key: Any) -> Optional[str]:
    normalized = _normalize_header(raw_key)
    for canonical, aliases in _COLUMN_ALIASES.items():
        if normalized in aliases:
            return canonical
    return None


class PadronService:
    def __init__(self):
        self.db = get_supabase_service()

    def get_consorcistas(self, search: Optional[str] = None) -> List[Dict[str, Any]]:
        query = self.db.client.table("consorcistas").select("*")
        if search:
            safe_search = _sanitize_search(search)
            if safe_search:
                query = query.or_(
                    f"nombre.ilike.%{safe_search}%,"
                    f"apellido.ilike.%{safe_search}%,"
                    f"cuit.ilike.%{safe_search}%"
                )
        result = query.order("apellido", desc=False).execute()
        return result.data

    def create_consorcista(self, data: Dict[str, Any]) -> Dict[str, Any]:
        result = self.db.client.table("consorcistas").insert(data).execute()
        return result.data[0] if result.data else {}

    def import_consorcistas(self, filename: str, content: bytes) -> Dict[str, Any]:
        rows = self._parse_rows(filename, content)

        processed = 0
        upserted = 0
        skipped = 0
        errors: List[Dict[str, Any]] = []

        for row_number, row in rows:
            processed += 1
            try:
                payload = self._build_consorcista_payload(row)
                if not payload:
                    skipped += 1
                    errors.append(
                        {
                            "row": row_number,
                            "error": "Fila sin datos mapeables para importacion",
                        }
                    )
                    continue

                if not payload.get("nombre") or not payload.get("apellido"):
                    skipped += 1
                    errors.append(
                        {
                            "row": row_number,
                            "error": "Nombre y apellido son obligatorios",
                        }
                    )
                    continue

                if not payload.get("cuit"):
                    skipped += 1
                    errors.append(
                        {
                            "row": row_number,
                            "error": "CUIT/CUIL es obligatorio para upsert",
                        }
                    )
                    continue

                self.db.client.table("consorcistas").upsert(
                    payload,
                    on_conflict="cuit",
                    ignore_duplicates=False,
                ).execute()
                upserted += 1
            except Exception as exc:
                skipped += 1
                errors.append({"row": row_number, "error": str(exc)})

        return {
            "processed": processed,
            "upserted": upserted,
            "skipped": skipped,
            "errors": errors,
        }

    def _build_consorcista_payload(self, row: Dict[str, Any]) -> Dict[str, Any]:
        mapped: Dict[str, Any] = {}
        for key, value in row.items():
            canonical = _canonical_key(key)
            if not canonical:
                continue
            mapped[canonical] = value

        payload: Dict[str, Any] = {
            "nombre": _normalize_text(mapped.get("nombre")),
            "apellido": _normalize_text(mapped.get("apellido")),
            "cuit": _normalize_cuit(mapped.get("cuit")),
            "representa_a": _normalize_text(mapped.get("representa_a")),
            "email": _normalize_text(mapped.get("email")),
            "telefono": _normalize_text(mapped.get("telefono")),
            "direccion_postal": _normalize_text(mapped.get("direccion_postal")),
            "activo": _normalize_active(mapped.get("activo")),
        }

        return {key: value for key, value in payload.items() if value is not None}

    def _parse_rows(
        self, filename: str, content: bytes
    ) -> List[Tuple[int, Dict[str, Any]]]:
        lower_name = filename.lower()
        if lower_name.endswith(".csv"):
            return self._parse_csv(content)
        if lower_name.endswith(".xlsx"):
            return self._parse_xlsx(content)
        if lower_name.endswith(".xls"):
            return self._parse_xls(content)
        raise ValueError("Formato no soportado. Usar CSV, XLS o XLSX")

    def _parse_csv(self, content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
        decoded = None
        for encoding in ("utf-8-sig", "latin-1"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue

        if decoded is None:
            raise ValueError("No se pudo decodificar CSV. Use UTF-8 o Latin-1")

        reader = csv.DictReader(io.StringIO(decoded))
        rows: List[Tuple[int, Dict[str, Any]]] = []
        for index, row in enumerate(reader, start=2):
            rows.append((index, dict(row)))
        return rows

    def _parse_xlsx(self, content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
        try:
            from openpyxl import load_workbook
        except Exception as exc:
            raise ValueError("Dependencia faltante para XLSX: openpyxl") from exc

        workbook = load_workbook(io.BytesIO(content), data_only=True)
        sheet = workbook.active
        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if not headers:
            return []

        headers_list = [str(value or "").strip() for value in headers]
        rows: List[Tuple[int, Dict[str, Any]]] = []
        for row_number, values in enumerate(rows_iter, start=2):
            row_dict = {
                headers_list[idx]: values[idx] if idx < len(values) else None
                for idx in range(len(headers_list))
            }
            rows.append((row_number, row_dict))
        return rows

    def _parse_xls(self, content: bytes) -> List[Tuple[int, Dict[str, Any]]]:
        try:
            import xlrd
        except Exception as exc:
            raise ValueError("Dependencia faltante para XLS: xlrd") from exc

        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return []

        headers = [
            str(sheet.cell_value(0, col) or "").strip() for col in range(sheet.ncols)
        ]
        rows: List[Tuple[int, Dict[str, Any]]] = []
        for row_idx in range(1, sheet.nrows):
            values: Iterable[Any] = [
                sheet.cell_value(row_idx, col) for col in range(sheet.ncols)
            ]
            row_dict = {headers[idx]: value for idx, value in enumerate(values)}
            rows.append((row_idx + 1, row_dict))
        return rows

    def get_pagos_by_consorcista(self, consorcista_id: UUID) -> List[Dict[str, Any]]:
        result = (
            self.db.client.table("cuotas_pagos")
            .select("*")
            .eq("consorcista_id", str(consorcista_id))
            .order("anio", desc=True)
            .execute()
        )
        return result.data

    def registrar_pago(self, pago_data: Dict[str, Any]) -> Dict[str, Any]:
        """Upsert a payment for a specific year."""
        result = (
            self.db.client.table("cuotas_pagos")
            .upsert(pago_data, on_conflict="consorcista_id,anio")
            .execute()
        )
        return result.data[0] if result.data else {}

    def get_deudores(self, anio: int) -> List[Dict[str, Any]]:
        """Find members who haven't paid a specific year."""
        members = self.get_consorcistas()
        pagos = (
            self.db.client.table("cuotas_pagos")
            .select("consorcista_id")
            .eq("anio", anio)
            .eq("estado", "pagado")
            .execute()
        )

        # Use a set for O(1) membership lookups instead of a list (O(n) per check)
        pagadores_ids: set[str] = {p["consorcista_id"] for p in pagos.data}

        return [m for m in members if str(m["id"]) not in pagadores_ids]


@lru_cache(maxsize=1)
def get_padron_service() -> PadronService:
    """Obtener instancia del servicio de padron (singleton)."""
    return PadronService()
